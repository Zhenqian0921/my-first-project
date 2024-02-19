import numpy as np
import openpyxl

def process_data():
    file_path = 'information.xlsx'  
    workbook = openpyxl.load_workbook(file_path)

# 选择包含节点信息的工作表
    nodes_sheet = workbook['Node']  

# 选择包含单元信息的工作表
    elements_sheet = workbook['Element']  

# 提取节点的坐标
    coordinates_x = []
    coordinates_y = []
    for row in nodes_sheet.iter_rows(min_row=2, values_only=True):#从第二行开始读取
        x = row[1]
        y = row[2]
        coordinates_x.append(x)
        coordinates_y.append(y)
    combined_array1 = np.column_stack((coordinates_x,coordinates_y))
    ID_node = combined_array1
#提取单元的连接关系，弹性模量和横截面积
    nodes1 = []
    nodes2 = []
    E = []
    A = []
    for row in elements_sheet.iter_rows(min_row=2, values_only=True):
        n1 = row[1]
        n2 = row[2]
        ela = row[3]
        area = row[4]
        nodes1.append(n1)
        nodes2.append(n2)
        E.append(ela)
        A.append(area)
    
    combined_array2 = np.column_stack((nodes1,nodes2,E,A))
    ID_element = combined_array2

#提取节点的边界条件    
    External_load_x = []
    External_load_y = []
    Fixed_situation_x = []
    Fixed_situation_y = []
    for row in nodes_sheet.iter_rows(min_row=2, values_only=True):
        load_x = row[3]
        load_y = row[4]
        Fix_x = row[5]
        Fix_y = row[6]
        External_load_x.append(load_x)
        External_load_y.append(load_y) 
        Fixed_situation_x.append(Fix_x)
        Fixed_situation_y.append(Fix_y)
    combined_array3 = np.column_stack((External_load_x, External_load_y , Fixed_situation_x, Fixed_situation_y))      
    BCs = combined_array3

#创建单元与节点的数组，计算单元的方向余弦和长度
    nodes_data = []
    for row in nodes_sheet.iter_rows(min_row=2, values_only=True):#从第二行开始读取
        x = row[1]
        y = row[2]
        nodes_data.append([x, y]) # 读取每个节点的坐标 
    elements_data = []
    for row in elements_sheet.iter_rows(min_row=2, values_only=True):
        node1 = row[1]
        node2 = row[2]
        elements_data.append([node1, node2]) # 读取每个单元与节点的连接关系
    elements_data = np.array(elements_data) - 1  # 减去1以匹配Python中的0-based索引
    c = []
    s = []
    L = []
    for i in range(len(elements_data)):
        node1, node2 = elements_data[i]
        x1, y1 = nodes_data[node1]
        x2, y2 = nodes_data[node2]
        Len = np.sqrt((x2 - x1)**2 + (y2 - y1)**2)
        cos = (x2 - x1) / Len
        sin = (y2 - y1) / Len
        c.append(cos)
        s.append(sin)
        L.append(Len)
    combined_array4 = np.column_stack((c, s, L))
    Information_element = combined_array4 
    return ID_node,ID_element,Information_element,BCs
    #print(ID_node,ID_element,Information_element)
#process_data()